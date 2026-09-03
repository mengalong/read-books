from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import fitz
from sqlalchemy import delete, select

from app.database import SessionLocal
from app.models import MaterialSegment, QuoteEntry, ResourceMaterial
from app.services.pdf_parser import extract_with_ocr, is_readable_page, normalize_text


logger = logging.getLogger(__name__)
MAX_SEGMENT_CHARS = 1_200
MAX_QUOTE_CHARS = 400
MIN_PDF_CHARS = 20


@dataclass(frozen=True)
class ParsedSegment:
    content: str
    page_number: int | None = None
    season_number: int | None = None
    episode_number: int | None = None
    start_ms: int | None = None
    end_ms: int | None = None
    scene_label: str | None = None
    speaker: str | None = None
    speaker_origin: str = "unknown"
    context: str | None = None


def normalized_quote_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    normalized = re.sub(r"\s+", "", normalized)
    return normalized.strip("\"'‘’“”")


def content_hash(value: str) -> str:
    return hashlib.sha256(normalized_quote_text(value).encode("utf-8")).hexdigest()


def _clean_dialogue(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value)
    value = re.sub(r"\{[^}]+\}", "", value)
    value = value.replace("\\N", "\n").replace("\\n", "\n")
    return normalize_text(value)


def _speaker_and_text(value: str) -> tuple[str | None, str]:
    text = _clean_dialogue(value)
    bracketed = re.match(r"^[【\[]([^】\]]{1,30})[】\]]\s*(.+)$", text, re.DOTALL)
    if bracketed:
        return bracketed.group(1).strip(), bracketed.group(2).strip()
    labeled = re.match(r"^([^：:\n]{1,30})[：:]\s*(.+)$", text, re.DOTALL)
    if labeled:
        speaker = labeled.group(1).strip()
        dialogue = labeled.group(2).strip()
        if not re.search(r"[，。！？!?]", speaker) and dialogue:
            return speaker, dialogue
    return None, text


def _parse_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def _timestamp_ms(value: str) -> int | None:
    match = re.fullmatch(r"(?:(\d{1,2}):)?(\d{1,2}):(\d{1,2})[,.](\d{1,3})", value.strip())
    if not match:
        return None
    hours = int(match.group(1) or 0)
    minutes = int(match.group(2))
    seconds = int(match.group(3))
    milliseconds = int(match.group(4).ljust(3, "0"))
    return ((hours * 60 + minutes) * 60 + seconds) * 1_000 + milliseconds


def _decode_text(path: Path) -> str:
    payload = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("文件编码无法识别，请转换为 UTF-8、UTF-16 或 GB18030 后重试")


def _episode_from_label(value: str | None) -> int | None:
    if not value:
        return None
    match = re.search(r"\d+", value)
    return int(match.group()) if match else None


def _split_long_text(text: str) -> list[str]:
    if len(text) <= MAX_SEGMENT_CHARS:
        return [text]
    return [
        text[start : start + MAX_SEGMENT_CHARS]
        for start in range(0, len(text), MAX_SEGMENT_CHARS)
    ]


def _parse_plain_text(material: ResourceMaterial) -> list[ParsedSegment]:
    text = _decode_text(Path(material.file_path))
    records: list[ParsedSegment] = []
    episode_number = _episode_from_label(material.episode_label)
    for raw_line in text.splitlines():
        line = _clean_dialogue(raw_line)
        if not line:
            continue
        speaker, dialogue = _speaker_and_text(line)
        for part in _split_long_text(dialogue):
            records.append(
                ParsedSegment(
                    content=part,
                    season_number=material.season_number,
                    episode_number=episode_number,
                    speaker=speaker,
                    speaker_origin="provided" if speaker else "unknown",
                )
            )
    return records


def _parse_pdf(material: ResourceMaterial) -> list[ParsedSegment]:
    pages: list[tuple[int, str]] = []
    with fitz.open(material.file_path) as document:
        if document.needs_pass and not document.authenticate(""):
            raise ValueError("PDF 受密码保护，无法读取其中的台词资料")
        for page_index, page in enumerate(document):
            pages.append((page_index + 1, page.get_text("text")))
        direct_is_usable = (
            sum(len(text) for _, text in pages) >= MIN_PDF_CHARS
            and any(is_readable_page(text) for _, text in pages)
        )
    if not direct_is_usable:
        pages = extract_with_ocr(material.file_path)

    records: list[ParsedSegment] = []
    episode_number = _episode_from_label(material.episode_label)
    for page_number, page_text in pages:
        for raw_line in page_text.splitlines():
            line = _clean_dialogue(raw_line)
            if not line:
                continue
            speaker, dialogue = _speaker_and_text(line)
            for part in _split_long_text(dialogue):
                records.append(
                    ParsedSegment(
                        content=part,
                        page_number=page_number,
                        season_number=material.season_number,
                        episode_number=episode_number,
                        speaker=speaker,
                        speaker_origin="provided" if speaker else "unknown",
                    )
                )
    return records


def _parse_srt_or_vtt(material: ResourceMaterial) -> list[ParsedSegment]:
    text = _decode_text(Path(material.file_path)).replace("\r\n", "\n")
    blocks = re.split(r"\n\s*\n", text)
    records: list[ParsedSegment] = []
    episode_number = _episode_from_label(material.episode_label)
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines or lines[0].upper().startswith(("WEBVTT", "NOTE", "STYLE", "REGION")):
            continue
        time_index = next((index for index, line in enumerate(lines) if "-->" in line), None)
        if time_index is None:
            continue
        time_parts = [part.strip().split()[0] for part in lines[time_index].split("-->", 1)]
        if len(time_parts) != 2:
            continue
        content = _clean_dialogue("\n".join(lines[time_index + 1 :]))
        if not content:
            continue
        speaker, dialogue = _speaker_and_text(content)
        records.append(
            ParsedSegment(
                content=dialogue,
                season_number=material.season_number,
                episode_number=episode_number,
                start_ms=_timestamp_ms(time_parts[0]),
                end_ms=_timestamp_ms(time_parts[1]),
                speaker=speaker,
                speaker_origin="provided" if speaker else "unknown",
            )
        )
    return records


def _parse_ass(material: ResourceMaterial) -> list[ParsedSegment]:
    text = _decode_text(Path(material.file_path))
    in_events = False
    fields: list[str] = []
    records: list[ParsedSegment] = []
    episode_number = _episode_from_label(material.episode_label)
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line.startswith("["):
            in_events = line.lower() == "[events]"
            continue
        if not in_events:
            continue
        if line.lower().startswith("format:"):
            fields = [item.strip().lower() for item in line.split(":", 1)[1].split(",")]
            continue
        if not line.lower().startswith("dialogue:") or not fields:
            continue
        values = line.split(":", 1)[1].split(",", len(fields) - 1)
        if len(values) != len(fields):
            continue
        row = dict(zip(fields, (value.strip() for value in values), strict=True))
        dialogue = _clean_dialogue(row.get("text", ""))
        if not dialogue:
            continue
        parsed_speaker, parsed_text = _speaker_and_text(dialogue)
        speaker = row.get("name") or parsed_speaker or None
        records.append(
            ParsedSegment(
                content=parsed_text,
                season_number=material.season_number,
                episode_number=episode_number,
                start_ms=_timestamp_ms(row.get("start", "")),
                end_ms=_timestamp_ms(row.get("end", "")),
                speaker=speaker,
                speaker_origin="provided" if speaker else "unknown",
            )
        )
    return records


FIELD_ALIASES = {
    "quote": ("台词", "quote", "dialogue", "text"),
    "speaker": ("角色", "speaker", "character", "name"),
    "season": ("季", "season"),
    "episode": ("集", "episode"),
    "start": ("开始时间", "start", "start_time"),
    "end": ("结束时间", "end", "end_time"),
    "scene": ("场景", "scene"),
    "context": ("上下文", "context"),
}


def _row_value(row: dict[str, Any], field: str) -> Any:
    lowered = {str(key).strip().lower(): value for key, value in row.items() if key is not None}
    for alias in FIELD_ALIASES[field]:
        if alias.lower() in lowered and lowered[alias.lower()] not in (None, ""):
            return lowered[alias.lower()]
    return None


def _rows_to_segments(material: ResourceMaterial, rows: list[dict[str, Any]]) -> list[ParsedSegment]:
    records: list[ParsedSegment] = []
    for position, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row.values()):
            continue
        quote = _clean_dialogue(str(_row_value(row, "quote") or ""))
        speaker = str(_row_value(row, "speaker") or "").strip()
        if not quote or not speaker:
            raise ValueError(f"台词表第 {position} 行缺少台词或角色")
        records.append(
            ParsedSegment(
                content=quote,
                season_number=_parse_int(_row_value(row, "season")) or material.season_number,
                episode_number=_parse_int(_row_value(row, "episode"))
                or _episode_from_label(material.episode_label),
                start_ms=_timestamp_ms(str(_row_value(row, "start") or "")),
                end_ms=_timestamp_ms(str(_row_value(row, "end") or "")),
                scene_label=str(_row_value(row, "scene") or "").strip() or None,
                speaker=speaker,
                speaker_origin="provided",
                context=str(_row_value(row, "context") or "").strip() or None,
            )
        )
    return records


def _parse_csv(material: ResourceMaterial) -> list[ParsedSegment]:
    text = _decode_text(Path(material.file_path))
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("台词表缺少表头")
    return _rows_to_segments(material, list(reader))


def _parse_xlsx(material: ResourceMaterial) -> list[ParsedSegment]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("当前环境缺少 XLSX 解析依赖，请重新安装后端依赖") from exc
    workbook = load_workbook(material.file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError("台词表缺少表头")
    payload = [dict(zip(headers, values, strict=False)) for values in rows]
    workbook.close()
    return _rows_to_segments(material, payload)


def parse_material_file(material: ResourceMaterial) -> list[ParsedSegment]:
    if material.file_format == "pdf":
        return _parse_pdf(material)
    if material.file_format == "txt":
        return _parse_plain_text(material)
    if material.file_format in {"srt", "vtt"}:
        return _parse_srt_or_vtt(material)
    if material.file_format == "ass":
        return _parse_ass(material)
    if material.file_format == "csv":
        return _parse_csv(material)
    if material.file_format == "xlsx":
        return _parse_xlsx(material)
    raise ValueError("不支持的可信资料格式")


def _context_for(records: list[ParsedSegment], index: int) -> str | None:
    if records[index].context:
        return records[index].context
    window = records[max(0, index - 1) : min(len(records), index + 2)]
    context = "\n".join(item.content for item in window if item.content)
    return context[:1_000] or None


def _is_quote_candidate(value: str) -> bool:
    normalized = normalized_quote_text(value)
    if len(normalized) < 2 or len(value) > MAX_QUOTE_CHARS:
        return False
    if re.fullmatch(r"[♪♫♬\s\W]+", value):
        return False
    return True


def parse_material_document(material_id: str) -> None:
    with SessionLocal() as db:
        material = db.get(ResourceMaterial, material_id)
        if material is None:
            return
        material.parse_status = "processing"
        material.error_message = None
        db.commit()

        try:
            records = parse_material_file(material)
            if not records:
                raise ValueError("资料中没有识别出可用文本")

            db.execute(delete(QuoteEntry).where(QuoteEntry.material_id == material.id))
            db.execute(delete(MaterialSegment).where(MaterialSegment.material_id == material.id))
            quote_count = 0
            pending_count = 0
            for sequence, record in enumerate(records, start=1):
                segment = MaterialSegment(
                    book_id=material.book_id,
                    material_id=material.id,
                    sequence=sequence,
                    content=record.content,
                    content_hash=content_hash(record.content),
                    page_number=record.page_number,
                    season_number=record.season_number,
                    episode_number=record.episode_number,
                    start_ms=record.start_ms,
                    end_ms=record.end_ms,
                    scene_label=record.scene_label,
                    speaker=record.speaker,
                    speaker_origin=record.speaker_origin,
                )
                db.add(segment)
                db.flush()
                if not _is_quote_candidate(record.content):
                    continue
                confirmed = record.speaker_origin == "provided"
                quote = QuoteEntry(
                    book_id=material.book_id,
                    material_id=material.id,
                    source_segment_ids=[segment.id],
                    quote_text=record.content,
                    normalized_text=normalized_quote_text(record.content),
                    content_hash=content_hash(record.content),
                    speaker=record.speaker,
                    speaker_origin=record.speaker_origin,
                    context=_context_for(records, sequence - 1),
                    season_number=record.season_number,
                    episode_number=record.episode_number,
                    start_ms=record.start_ms,
                    end_ms=record.end_ms,
                    page_number=record.page_number,
                    review_status="confirmed" if confirmed else "pending",
                    enabled_for_generation=confirmed,
                )
                db.add(quote)
                quote_count += 1
                pending_count += 0 if confirmed else 1

            material.segment_count = len(records)
            material.quote_count = quote_count
            material.parse_status = "needs_review" if pending_count else "completed"
            db.commit()
        except Exception as exc:
            logger.exception("可信资料解析失败: %s", material_id)
            db.rollback()
            failed = db.get(ResourceMaterial, material_id)
            if failed:
                failed.parse_status = "failed"
                failed.error_message = str(exc) or "可信资料解析失败"
                db.commit()


def recover_material_tasks() -> list[str]:
    with SessionLocal() as db:
        material_ids = list(
            db.scalars(
                select(ResourceMaterial.id).where(
                    ResourceMaterial.parse_status.in_(["pending", "processing"])
                )
            ).all()
        )
        if material_ids:
            db.query(ResourceMaterial).filter(ResourceMaterial.id.in_(material_ids)).update(
                {ResourceMaterial.parse_status: "pending"}, synchronize_session=False
            )
            db.commit()
        return material_ids
